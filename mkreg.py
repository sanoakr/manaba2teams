from csv import reader
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

csvFile = "履修登録.csv"
todayStr = datetime.now().strftime("%Y%m%d")
xlsxDayFile = csvFile.replace(".csv", f"{todayStr}.xlsx")
xlsxFile = csvFile.replace(".csv",".xlsx")

with open(csvFile) as csvf:
    # excel book
    wb = Workbook()
    ws = wb.active
    # csv read
    data = reader(csvf, delimiter=',')
    for row in data:
        ws.append(row)
    # delete 1 col
    ws.delete_cols(1)
    # delete 1,3,last row
    ws.delete_rows(1)
    ws.delete_rows(2) # original 3rd row!!
    ws.delete_rows(ws.max_row)
    # table
    all = f'A1:G{ws.max_row}'
    table = Table(displayName='record', ref=all)
    table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium8', showRowStripes=True)
    ws.add_table(table)

    # excel write
    wb.save(xlsxDayFile)
    wb.save(xlsxFile)
